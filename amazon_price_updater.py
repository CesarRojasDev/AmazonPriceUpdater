import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException, NoSuchElementException
from openpyxl.utils import column_index_from_string
from colorama import init, Fore, Style
import time

# Inicializar colorama
init(autoreset=True)

def get_column_index(column_letter):
    return column_index_from_string(column_letter)

def display_menu():
    print(Fore.YELLOW + "****************************************")
    print(Fore.YELLOW + "********* " + Fore.CYAN + "AUTOMATIZACIÓN DE PRECIOS" + Fore.YELLOW + " *********")
    print(Fore.YELLOW + "****************************************")
    print(Fore.YELLOW + "* " + Fore.GREEN + "1. Ejecutar el script" + Fore.YELLOW + " *")
    print(Fore.YELLOW + "* " + Fore.RED + "2. Salir" + Fore.YELLOW + " *")
    print(Fore.YELLOW + "****************************************")

def main(archivo, columna_precio, columna_link, columna_stock):
    driver = webdriver.Chrome()

    archivo_excel = f"{archivo}.xlsx"
    wb = openpyxl.load_workbook(archivo_excel)
    hoja = wb.active

    columna_precio_idx = get_column_index(columna_precio)
    columna_link_idx = get_column_index(columna_link)
    columna_stock_idx = get_column_index(columna_stock)

    first_product = True

    for fila in hoja.iter_rows(min_row=2, max_row=hoja.max_row, min_col=columna_link_idx, max_col=columna_link_idx):
        enlace = fila[0].value

        if enlace and enlace.startswith("https://www.amazon"):
            try:
                driver.get(enlace)
                WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.ID, "productTitle"))
                )

                if first_product:
                    time.sleep(10)
                    first_product = False

                disponible = False
                try:
                    driver.find_element(By.ID, "add-to-cart-button")
                    disponible = True
                except NoSuchElementException:
                    pass

                if disponible:
                    try:
                        precio_entero_element = driver.find_element(By.XPATH, "//span[contains(@class, 'a-price-whole')]")
                        precio_entero = precio_entero_element.text.strip().replace(",", "")
                        precio_decimal_element = driver.find_element(By.XPATH, "//span[contains(@class, 'a-price-fraction')]")
                        precio_decimal = precio_decimal_element.text.strip()
                        precio = float(f"{precio_entero}.{precio_decimal}")

                        precio_redondeado = round(precio) + 1
                        hoja.cell(row=fila[0].row, column=columna_stock_idx).value = "4"
                    except (NoSuchElementException, ValueError) as e:
                        print(Fore.RED + "Error al obtener el precio:", e)
                        precio_redondeado = 0
                        hoja.cell(row=fila[0].row, column=columna_stock_idx).value = "0"
                else:
                    precio_redondeado = 0
                    hoja.cell(row=fila[0].row, column=columna_stock_idx).value = "0"

            except (TimeoutException, WebDriverException) as e:
                print(Fore.RED + "Error:", e)
                precio_redondeado = 0
                hoja.cell(row=fila[0].row, column=columna_stock_idx).value = "0"
                hoja.cell(row=fila[0].row, column=columna_precio_idx).value = 0
                continue

            hoja.cell(row=fila[0].row, column=columna_precio_idx).value = precio_redondeado

            time.sleep(5)

    driver.quit()
    wb.save(archivo_excel)

if __name__ == "__main__":
    while True:
        display_menu()
        choice = input(Fore.CYAN + "Seleccione una opción: ")

        if choice == '1':
            print(Fore.YELLOW + "****************************************")
            print(Fore.YELLOW + "********** " + Fore.CYAN + "EJECUTAR EL SCRIPT" + Fore.YELLOW + " **********")
            print(Fore.YELLOW + "****************************************")
            archivo = input(Fore.GREEN + "Ingrese el nombre del archivo Excel (sin la extensión): ")
            columna_precio = input(Fore.GREEN + "Ingrese la letra de la columna del precio: ")
            columna_link = input(Fore.GREEN + "Ingrese la letra de la columna del enlace: ")
            columna_stock = input(Fore.GREEN + "Ingrese la letra de la columna del stock: ")
            print(Fore.YELLOW + "****************************************")
            main(archivo, columna_precio, columna_link, columna_stock)
            print(Fore.YELLOW + "****************************************")
            print(Fore.GREEN + "** El script se ha ejecutado exitosamente. **")
            print(Fore.YELLOW + "****************************************")
        elif choice == '2':
            print(Fore.RED + "Saliendo del programa...")
            break
        else:
            print(Fore.RED + "Opción no válida. Por favor, intente de nuevo.")
